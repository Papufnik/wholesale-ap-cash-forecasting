import os
import json
from datetime import datetime, date, timedelta

"""
Wholesale AP Cash-Need Forecast
================================
(Anonymized portfolio version of a script running in production for a real
seasonal retail business. Business name, dollar figures, and file paths
below are illustrative, not real.)

WHY THIS EXISTS: the business orders inventory from a wholesale marketplace
on net terms. The marketplace's dashboard shows the total owed, but not a
reliable per-order charge date -- an earlier internal tracker estimated due
dates as "order date + 60 days," which turned out to be wrong on a large
share of orders (real charge dates ranged 60-119+ days out depending on
ship date and the vendor's own billing cycle). That gap produced false
"overdue" alarms that didn't match the vendor's own records.

WHAT THIS DOES: once each order's real charge date is available, this script
turns that into a forward cash-need calendar -- how much needs to be in the
bank, and by when, to cover every remaining obligation.

UNDATED ORDERS: some orders are charged automatically to a card with no
fixed date shown anywhere. Rather than guessing a date for these, they are
carried as a separate standing reserve on top of the dated schedule --
never assume a date that isn't there.

Current cash is read live from the business's own ledger export.
"""

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "sample_data")
SCHEDULE_JSON = os.path.join(DATA_DIR, "vendor_ap_schedule.json")
LEDGER_XLSX = os.path.join(DATA_DIR, "operating_ledger.xlsx")


def _load_operating_cash():
    import openpyxl
    wb = openpyxl.load_workbook(LEDGER_XLSX, data_only=True)
    ws = wb.active
    bal, dt = None, None
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[5] is not None:
            bal, dt = row[5], row[0]
    return bal, dt


def _week_start(d):
    return d - timedelta(days=d.weekday())  # Monday-anchored


def build_cash_calendar():
    print(f"=== Wholesale AP Cash-Need Calendar [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ===")

    with open(SCHEDULE_JSON, encoding="utf-8") as f:
        schedule = json.load(f)

    rows = schedule.get("dueDateDetail") or []
    dated = [r for r in rows if r.get("dueDate")]
    undated = [r for r in rows if not r.get("dueDate")]

    today = date.today()

    by_day = {}
    for r in dated:
        d = datetime.strptime(r["dueDate"], "%Y-%m-%d").date()
        by_day.setdefault(d, []).append(r)

    daily = []
    for d in sorted(by_day):
        rows_d = by_day[d]
        daily.append({
            "date": d.strftime("%Y-%m-%d"),
            "amount": round(sum(x["amount"] for x in rows_d), 2),
            "orders": len(rows_d),
            "daysFromToday": (d - today).days,
        })

    by_week = {}
    for d, rows_d in by_day.items():
        wk = _week_start(d)
        by_week.setdefault(wk, []).extend(rows_d)

    weekly = []
    running_total = 0.0
    for wk in sorted(by_week):
        rows_w = by_week[wk]
        wk_amount = round(sum(x["amount"] for x in rows_w), 2)
        running_total = round(running_total + wk_amount, 2)
        weekly.append({
            "weekOf": wk.strftime("%Y-%m-%d"),
            "amountDueThisWeek": wk_amount,
            "cumulativeReserveNeeded": running_total,
        })

    undated_total = round(sum(x["amount"] for x in undated), 2)

    op_cash, op_asof = _load_operating_cash()
    total_faire_remaining = schedule.get("totalRemaining")

    payload = {
        "asOfDate": today.strftime("%Y-%m-%d"),
        "totalRemainingThisSeason": total_faire_remaining,
        "scheduledTotal": round(sum(d["amount"] for d in daily), 2),
        "undatedAutopayTotal": undated_total,
        "currentOperatingCash": op_cash,
        "operatingCashAsOf": op_asof.strftime("%Y-%m-%d") if op_asof else None,
        "daily": daily,
        "weekly": weekly,
    }

    print(f"  Total remaining this season: ${total_faire_remaining:,.2f} "
          f"(${payload['scheduledTotal']:,.2f} dated + ${undated_total:,.2f} undated/autopay)")
    if op_cash is not None:
        print(f"  Current operating cash: ${op_cash:,.2f} as of {op_asof.strftime('%Y-%m-%d')}")
    for w in weekly:
        print(f"    Week of {w['weekOf']}: ${w['amountDueThisWeek']:,.2f} due, "
              f"cumulative ${w['cumulativeReserveNeeded']:,.2f}")
    return payload


if __name__ == "__main__":
    build_cash_calendar()
